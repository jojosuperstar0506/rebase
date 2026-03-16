"""Formatted Excel output generator.

Produces a 2-sheet Excel workbook:
- Sheet 1: 数据源_合并 — merged raw data, filter-ready for Feishu Bitable import
- Sheet 2: 产品结构分析 — 11 continuous analysis sections
"""

from __future__ import annotations

from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from ..models.schemas import AnalysisResult
except ImportError:
    from models.schemas import AnalysisResult


# --- Style constants ---
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="2F5496")
NORMAL_FONT = Font(name="Microsoft YaHei", size=10)
NUMBER_FONT = Font(name="Microsoft YaHei", size=10)
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

GRADE_COLORS = {
    "A·明星款": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "B·稳定款": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "C·观察款": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "D·淘汰候选": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


def write_analysis_excel(
    merged_df: pd.DataFrame,
    result: AnalysisResult,
    output_path: Optional[str] = None,
) -> BytesIO:
    """Generate formatted Excel workbook with 2 sheets.

    Args:
        merged_df: The merged dataset from data_merger
        result: The AnalysisResult from analyzer
        output_path: Optional file path to save (also returns BytesIO)

    Returns:
        BytesIO buffer containing the Excel file.
    """
    wb = Workbook()

    # --- Sheet 1: 数据源_合并 ---
    ws1 = wb.active
    ws1.title = "数据源_合并"
    _write_data_source_sheet(ws1, merged_df)

    # --- Sheet 2: 产品结构分析 ---
    ws2 = wb.create_sheet("产品结构分析")
    _write_analysis_sheet(ws2, result)

    # Save
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    if output_path:
        with open(output_path, "wb") as f:
            f.write(buffer.getvalue())
        buffer.seek(0)

    return buffer


def _write_data_source_sheet(ws, df: pd.DataFrame) -> None:
    """Write merged data as a filterable table."""
    # Select and order columns for export
    export_cols = [
        "sku_code", "product_name", "material", "bag_type", "price_tier",
        "retail_price", "cost_estimate", "sales_volume", "return_volume",
        "net_volume", "return_rate", "net_revenue", "estimated_margin",
        "margin_rate", "avg_monthly_sales", "efficiency_rating",
        "current_stock", "stock_to_sales_ratio", "inventory_status",
    ]
    available_cols = [c for c in export_cols if c in df.columns]

    col_labels = {
        "sku_code": "款式编码",
        "product_name": "商品名称",
        "material": "材质",
        "bag_type": "包型",
        "price_tier": "价格带",
        "retail_price": "售价",
        "cost_estimate": "估算成本",
        "sales_volume": "销售数量",
        "return_volume": "退货数量",
        "net_volume": "净销量",
        "return_rate": "退货率",
        "net_revenue": "净销售额",
        "estimated_margin": "估算毛利",
        "margin_rate": "毛利率",
        "avg_monthly_sales": "月均净销量",
        "efficiency_rating": "效率评级",
        "current_stock": "现库存",
        "stock_to_sales_ratio": "库销比(月)",
        "inventory_status": "库存状态",
    }

    # Write headers
    for col_idx, col in enumerate(available_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_labels.get(col, col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Write data
    pct_cols = {"return_rate", "margin_rate", "revenue_share"}
    money_cols = {"retail_price", "cost_estimate", "net_revenue", "estimated_margin"}

    for row_idx, (_, row) in enumerate(df[available_cols].iterrows(), 2):
        for col_idx, col in enumerate(available_cols, 1):
            val = row[col]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

            # Format percentages
            if col in pct_cols and isinstance(val, (int, float)):
                cell.number_format = "0.0%"

            # Format money
            if col in money_cols and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"

            # Color-code efficiency rating
            if col == "efficiency_rating" and val in GRADE_COLORS:
                cell.fill = GRADE_COLORS[val]

            # Color-code inventory status
            if col == "inventory_status":
                if val == "健康":
                    cell.fill = GREEN_FILL
                elif val == "偏高":
                    cell.fill = YELLOW_FILL
                elif val in ("严重积压", "零动销"):
                    cell.fill = RED_FILL

    # Auto-fit column widths (approximate)
    for col_idx, col in enumerate(available_cols, 1):
        label = col_labels.get(col, col)
        width = max(len(label) * 2, 12)  # Chinese chars ~2x width
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 25)

    # Add auto-filter
    if available_cols:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(available_cols))}{len(df) + 1}"


def _write_analysis_sheet(ws, result: AnalysisResult) -> None:
    """Write the 11-section analysis document."""
    row = 1

    # --- Section A: KPI总览 ---
    row = _write_section_header(ws, row, "A. KPI总览")
    s = result.summary
    kpis = [
        ("总SKU数", s.total_skus, None),
        ("活跃SKU数", s.active_skus, None),
        ("总销售量", s.total_sales_volume, "#,##0"),
        ("总退货量", s.total_return_volume, "#,##0"),
        ("整体退货率", s.overall_return_rate, "0.0%"),
        ("净销售额", s.total_net_revenue, "¥#,##0.00"),
        ("估算总毛利", s.total_estimated_margin, "¥#,##0.00"),
        ("整体毛利率", s.overall_margin_rate, "0.0%"),
        ("现库存总量", s.total_current_stock, "#,##0"),
        ("A级明星款", s.grade_a_count, None),
        ("D级淘汰候选", s.grade_d_count, None),
        ("头部SKU", f"{s.top_sku_code} (¥{s.top_sku_revenue:,.0f})", None),
    ]
    for label, value, fmt in kpis:
        ws.cell(row=row, column=1, value=label).font = Font(name="Microsoft YaHei", size=10, bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = NUMBER_FONT
        if fmt:
            cell.number_format = fmt
        row += 1
    row += 1

    # --- Sections B-D: Dimension breakdowns ---
    for dim_name in ["材质", "包型", "价格带"]:
        section_label = {"材质": "B", "包型": "C", "价格带": "D"}[dim_name]
        row = _write_section_header(ws, row, f"{section_label}. {dim_name}分析")

        items = [b for b in result.dimension_breakdowns if b.dimension_name == dim_name]
        if items:
            headers = [dim_name, "SKU数", "占比", "销量", "退货", "退货率",
                       "净销量", "净销售额", "占比", "毛利", "毛利率", "均价"]
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
            row += 1

            for item in items:
                vals = [
                    item.dimension_value, item.sku_count, item.sku_share,
                    item.sales_volume, item.return_volume, item.return_rate,
                    item.net_volume, item.net_revenue, item.revenue_share,
                    item.estimated_margin, item.margin_rate, item.avg_price,
                ]
                for ci, val in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.font = NORMAL_FONT
                    cell.border = THIN_BORDER
                    # Format percentages
                    if ci in (3, 6, 9, 11):
                        cell.number_format = "0.0%"
                    elif ci in (8, 10, 12):
                        cell.number_format = "#,##0.00"
                    # Highlight high return rates
                    if ci == 6 and isinstance(val, (int, float)):
                        if val > 0.40:
                            cell.fill = RED_FILL
                        elif val < 0.25:
                            cell.fill = GREEN_FILL
                row += 1
        row += 1

    # --- Section E: Efficiency grades ---
    row = _write_section_header(ws, row, "E. SKU效率评级分布")
    headers = ["评级", "标准", "SKU数", "占比", "净销售额", "营收占比", "平均退货率"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for grade in result.efficiency_grades:
        vals = [
            grade.rating, grade.criteria, grade.sku_count,
            grade.sku_share, grade.total_net_revenue,
            grade.revenue_share, grade.avg_return_rate,
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if ci in (4, 6, 7):
                cell.number_format = "0.0%"
            elif ci == 5:
                cell.number_format = "¥#,##0.00"
        if grade.rating in GRADE_COLORS:
            for ci in range(1, 8):
                ws.cell(row=row, column=ci).fill = GRADE_COLORS[grade.rating]
        row += 1
    row += 1

    # --- Section F: Inventory health ---
    if result.inventory_health:
        row = _write_section_header(ws, row, "F. 库存健康度")
        headers = ["款式编码", "商品名称", "材质", "现库存", "周日均销量",
                    "剩余天数", "库销比(历史)", "库销比(近期)", "状态"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        row += 1

        for item in result.inventory_health[:50]:  # Top 50
            vals = [
                item.sku_code, item.product_name, item.material,
                item.current_stock, item.weekly_daily_sales,
                item.days_remaining, item.stock_months_historical,
                item.stock_months_recent, item.status,
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
            # Color status
            status_cell = ws.cell(row=row, column=9)
            if item.status == "健康":
                status_cell.fill = GREEN_FILL
            elif item.status == "偏高":
                status_cell.fill = YELLOW_FILL
            elif item.status in ("严重积压", "零动销"):
                status_cell.fill = RED_FILL
            row += 1
        row += 1

    # --- Section G: Purchase recommendations ---
    if result.purchase_recommendations:
        row = _write_section_header(ws, row, "G. 采购建议")
        headers = ["材质", "目标库销比(月)", "现库存", "月均净销量",
                    "建议补货量", "预估补货成本", "说明"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        row += 1

        for rec in result.purchase_recommendations:
            vals = [
                rec.category, rec.target_wos_months, rec.current_stock,
                rec.avg_monthly_sales, rec.recommended_order_qty,
                rec.estimated_order_cost, rec.rationale,
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                if ci == 6:
                    cell.number_format = "¥#,##0.00"
            row += 1
        row += 1

    # --- Section H: Key insights ---
    row = _write_section_header(ws, row, "H. 关键洞察与行动建议")
    for i, insight in enumerate(result.key_insights, 1):
        ws.cell(row=row, column=1, value=f"{i}.").font = Font(
            name="Microsoft YaHei", size=10, bold=True
        )
        ws.cell(row=row, column=2, value=insight).font = NORMAL_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        row += 1

    # Set column widths
    widths = [15, 20, 15, 12, 12, 12, 15, 15, 12, 15, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_section_header(ws, row: int, title: str) -> int:
    """Write a section header row and return the next row number."""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    return row + 1
